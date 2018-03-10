from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import math

# ================= initialize =============================
WB = load_workbook(filename='ResponseSpectrum.xlsx')

WS = WB['Sheet1']


def get_data_byexcel(min_row, min_col, max_row, max_col):
    data = np.array([])
    for row in WS.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            data = np.append(data, [cell.value])
    return np.reshape(data, (max_row - min_row + 1, max_col - min_col + 1))


ACCELERATIONS = get_data_byexcel(min_row=3, min_col=2, max_row=82, max_col=6)
PERIODS = get_data_byexcel(min_row=3, min_col=1, max_row=82, max_col=1)


# ================= hw1_2_a: 取加速度最大值 =============================
pga = np.amax(ACCELERATIONS, axis=0)
print("PGA:", pga)


# ================= hw1_2_b:正規化加速度後繪圖 =============================
def scale_linear_bycolumn(raw_data, norm=100.0):
    benchmark = raw_data[0]
    return raw_data * norm / benchmark


normalize_acceleration = scale_linear_bycolumn(
    raw_data=ACCELERATIONS, norm=0.33)

plt.plot(PERIODS, normalize_acceleration)
plt.show()


# ================= hw1_2_c:解釋 αy、Fu、1.4 =============================
# αy：αy is defined as the first yield seismic force amplification factor that is dependent on the structure types and design method.
# Fu：the seismic force reduction factor. This is based on the equal displacement principle between the elastic and the EPP systems for the long period range and the equal energy principle for short periods.
# 1.4：the constant 1.4 means the over-strength factor between the ultimate and the first yield force.

# ================= hw1_2_d:繪圖 =============================
important_factor = 1
alpha_y = 1
sd1 = 0.48
sds = 0.78
sm1 = 0.55
sms = 0.88
ductility_capacity = 4.8

period0 = sd1 / sds

ra = 1 + (ductility_capacity - 1) / 1.5

for period in PERIODS:

    # sad
    if period <= 0.2 * period0:
        sad = sds * (0.4 + 3 * period / period0)
    elif 0.2 * period0 <= period <= period0:
        sad = sds
    elif period0 <= period <= 2.5 * period0:
        sad = sd1 / period
    elif period >= 2.5 * period0:
        sad = 0.4 * sds

    # fu
    if period >= period0:
        fu = ra
    elif 0.6 * period0 <= period <= period0:
        fu = math.sqrt(2 * ra - 1) + (ra - math.sqrt(2 * ra - 1)) * \
            (period - 0.6 * period0) / (0.4 * period0)
    elif 0.2 * period0 <= period <= 0.6 * period0:
        fu = math.sqrt(2 * ra - 1)
    elif period <= 0.2 * period0:
        fu = math.sqrt(2 * ra - 1) + (math.sqrt(2 * ra - 1) - 1) * \
            (period - 0.2 * period0) / (0.2 * period0)

    if expression:
        pass

    design_base_shear_factor = important_factor / \
        (1.4 * alpha_y) * design_modified_ratio
    max_base_shear_factor = important_factor / \
        (1.4 * alpha_y) * max_modified_ratio
    min_base_shear_factor = important_factor * \
        fu / (4.2 * alpha_y) * design_modified_ratio
