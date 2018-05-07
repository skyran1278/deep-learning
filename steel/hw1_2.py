import math

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

# ================= initialize =============================
WB = load_workbook(
    filename='Seismic Design of Steel Structures/ResponseSpectrum.xlsx')

WS = WB['Sheet1']


def get_data_byexcel(min_row, min_col, max_row, max_col):
    data = np.array([])
    for row in WS.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            data = np.append(data, [cell.value])
    return np.reshape(data, (max_row - min_row + 1, max_col - min_col + 1))


ACCELERATIONS = get_data_byexcel(min_row=3, min_col=2, max_row=82, max_col=6)
DISPLACEMENT_R3 = get_data_byexcel(
    min_row=3, min_col=9, max_row=82, max_col=13)
DISPLACEMENT_R5 = get_data_byexcel(
    min_row=3, min_col=16, max_row=82, max_col=20)
PERIODS = get_data_byexcel(min_row=3, min_col=1, max_row=82, max_col=1)


# ================= hw1_2_a: 取加速度最大值 =============================
def hw1_2_a():
    pga = np.amax(ACCELERATIONS, axis=0)
    # print('PGA: \n', pga)


hw1_2_a()


# ================= hw1_2_b:正規化加速度後繪圖 =============================


def scale_linear_bycolumn(raw_data, norm=100.0):
    benchmark = raw_data[0]
    return raw_data * norm / benchmark


def hw1_2_b():
    normalize_acceleration = scale_linear_bycolumn(
        raw_data=ACCELERATIONS, norm=0.33)

    # print(normalize_acceleration)
    plt.figure()
    plt.plot(PERIODS, normalize_acceleration)
    plt.legend(['TCU015', 'TCU029', 'TCU076', 'TCU082', 'TCU089'])

    return normalize_acceleration


NORMALIZE_ACCELERATION = hw1_2_b()


# ================= hw1_2_c:解釋 αy、Fu、1.4 =============================
# αy：αy is defined as the first yield seismic force amplification factor that is dependent on the structure types and design method.
# Fu：the seismic force reduction factor. This is based on the equal displacement principle between the elastic and the EPP systems for the long period range and the equal energy principle for short periods.
# 1.4：the constant 1.4 means the over-strength factor between the ultimate and the first yield force.


# ================= hw1_2_d:繪圖 =============================
def hw1_2_d():
    important_factor = 1
    alpha_y = 1
    sd1 = 0.48
    sds = 0.78
    sm1 = 0.55
    sms = 0.88
    ductility_capacity = 4.8

    period_0 = sd1 / sds
    period_0_m = sm1 / sms

    ra = 1 + (ductility_capacity - 1) / 1.5

    base_shear_factor = np.array([])
    quotients = np.array([])

    for period in PERIODS:

        # sad
        if period <= 0.2 * period_0:
            sad = sds * (0.4 + 3 * period / period_0)
        elif 0.2 * period_0 <= period <= period_0:
            sad = sds
        elif period_0 <= period <= 2.5 * period_0:
            sad = sd1 / period
        elif period > 2.5 * period_0:
            sad = 0.4 * sds

        # fu
        if period >= period_0:
            fu = ra
        elif 0.6 * period_0 <= period <= period_0:
            fu = math.sqrt(2 * ra - 1) + (ra - math.sqrt(2 * ra - 1)) * \
                (period - 0.6 * period_0) / (0.4 * period_0)
        elif 0.2 * period_0 <= period <= 0.6 * period_0:
            fu = math.sqrt(2 * ra - 1)
        elif period <= 0.2 * period_0:
            fu = math.sqrt(2 * ra - 1) + (math.sqrt(2 * ra - 1) - 1) * \
                (period - 0.2 * period_0) / (0.2 * period_0)

        # design_modified_ratio
        if sad / fu <= 0.3:
            design_modified_ratio = sad / fu
        elif 0.3 <= sad / fu <= 0.8:
            design_modified_ratio = 0.52 * sad / fu + 0.144
        elif sad / fu > 0.8:
            design_modified_ratio = 0.7 * sad / fu

        # sam
        if period <= 0.2 * period_0_m:
            sam = sms * (0.4 + 3 * period / period_0_m)
        elif 0.2 * period_0_m <= period <= period_0_m:
            sam = sms
        elif period_0_m <= period <= 2.5 * period_0_m:
            sam = sm1 / period
        elif period > 2.5 * period_0_m:
            sam = 0.4 * sms

        # fu_m
        if period >= period_0_m:
            fu_m = ductility_capacity
        elif 0.6 * period_0_m <= period <= period_0_m:
            fu_m = math.sqrt(2 * ductility_capacity - 1) + (ductility_capacity - math.sqrt(2 * ductility_capacity - 1)) * \
                (period - 0.6 * period_0_m) / (0.4 * period_0_m)
        elif 0.2 * period_0_m <= period <= 0.6 * period_0_m:
            fu_m = math.sqrt(2 * ductility_capacity - 1)
        elif period <= 0.2 * period_0_m:
            fu_m = math.sqrt(2 * ductility_capacity - 1) + (math.sqrt(2 * ductility_capacity - 1) - 1) * \
                (period - 0.2 * period_0_m) / (0.2 * period_0_m)

        if sam / fu_m <= 0.3:
            max_modified_ratio = sam / fu_m
        elif 0.3 <= sam / fu_m <= 0.8:
            max_modified_ratio = 0.52 * sam / fu_m + 0.144
        elif sam / fu_m > 0.8:
            max_modified_ratio = 0.7 * sam / fu_m

        design_base_shear_factor = important_factor / \
            (1.4 * alpha_y) * design_modified_ratio
        max_base_shear_factor = important_factor / \
            (1.4 * alpha_y) * max_modified_ratio
        min_base_shear_factor = important_factor * \
            fu / (4.2 * alpha_y) * design_modified_ratio

        quotients = np.append(quotients, [fu])
        base_shear_factor = np.append(base_shear_factor, [
            design_base_shear_factor, max_base_shear_factor, min_base_shear_factor])

    base_shear_factor = np.reshape(base_shear_factor, (80, 3))

    plt.figure()
    plt.plot(PERIODS, base_shear_factor)
    plt.legend(['V/W', 'V*/W', 'VM/W'])

    return base_shear_factor, 1.4 * alpha_y * quotients


BASE_SHEAR_FACTOR, QUOTIENTS = hw1_2_d()


# ================= hw1_2_e: =============================
def hw1_2_e():
    design_base_shear_factor = BASE_SHEAR_FACTOR[:, 0]
    factor = np.mean(NORMALIZE_ACCELERATION, axis=1) / design_base_shear_factor

    plt.figure()
    plt.plot(PERIODS, factor)
    plt.plot(PERIODS, QUOTIENTS)
    plt.legend(['averaged acceleration response spectra divide V/W', '1.4*αy*Fu'])


hw1_2_e()


# ================= hw1_2_f: =============================
def hw1_2_f():
    omega = 2 * math.pi / PERIODS
    average_acceleration = np.mean(NORMALIZE_ACCELERATION, axis=1)
    average_acceleration = np.reshape(average_acceleration, (80, 1))

    plt.figure()
    plt.plot(PERIODS, NORMALIZE_ACCELERATION * 9.81 / (omega ** 2))
    plt.plot(PERIODS, average_acceleration * 9.81 / (omega ** 2))
    plt.legend(['TCU015', 'TCU029', 'TCU076', 'TCU082', 'TCU089', 'AVERAGE'])

    return NORMALIZE_ACCELERATION * 9.81 / (omega ** 2)


DISPLACEMENT_ELASTIC = hw1_2_f()


# ================= hw1_2_g: =============================
def hw1_2_g():
    cr3 = DISPLACEMENT_R3 / DISPLACEMENT_ELASTIC
    cr5 = DISPLACEMENT_R5 / DISPLACEMENT_ELASTIC

    plt.figure()
    plt.plot(PERIODS, cr3, linestyle='--')
    plt.plot(PERIODS, cr5, linestyle=':')
    plt.plot(PERIODS, np.ones((80, 1)))
    plt.show()


hw1_2_g()
