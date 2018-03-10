from openpyxl import load_workbook
import numpy as np

wb = load_workbook(filename='ResponseSpectrum.xlsx')

ws = wb['Sheet1']

acceleration = np.array([])

TCU015 = 0
TCU029 = 1
TCU076 = 2
TCU082 = 3
TCU089 = 4

for row in ws.iter_rows(min_row=3, min_col=2, max_row=82, max_col=6):
    for cell in row:
        acceleration = np.append(acceleration, [cell.value])

acceleration = np.reshape(acceleration, (80, 5))

normalize_acceleration = scale_linear_bycolumn(
    raw_data=acceleration, norm=0.33)

print(normalize_acceleration)


def scale_linear_bycolumn(raw_data, norm=100.0):
    benchmark = raw_data[0]
    return raw_data * norm / benchmark
