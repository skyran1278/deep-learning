from openpyxl import load_workbook
wb = load_workbook(filename='ResponseSpectrum.xlsx')
ws = wb['Sheet1']
TCU = []

TCU[0] = ws['B3':'B82']

for rows in cell_range:
    for cell in rows:
        print(cell.value)
